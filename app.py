# app.py
import os
import threading
import time
import random
import smtplib
from email.mime.text import MIMEText
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import unquote

from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_bcrypt import Bcrypt
from pymongo import MongoClient
from dotenv import load_dotenv
import openpyxl

# -------------------------------
# Load env
# -------------------------------
load_dotenv()

app = Flask(__name__)
CORS(app)
bcrypt = Bcrypt(app)

# -------------------------------
# MongoDB setup
# -------------------------------
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/")
client = MongoClient(MONGO_URI)
db = client.get_database("coinsora")
users = db["users"]
pending_users = db["pending_users"]

# -------------------------------
# Email setup (Gmail SMTP)
# -------------------------------
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")

# -------------------------------
# Excel config
# -------------------------------
BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "storage.xlsx"

# -------------------------------
# Defaults
# -------------------------------
DEFAULT_PROFILE_PIC = "https://i.pravatar.cc/150?img=12"

# -------------------------------
# Helper: Send Email OTP
# -------------------------------
def send_email_otp(recipient, otp):
    subject = "Your Coinsora OTP Verification Code"
    body = f"Your OTP is {otp}. It expires in 5 minutes."

    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        server.send_message(msg)

# -------------------------------
# Cleanup expired OTPs (background)
# -------------------------------
def clean_expired_otps():
    while True:
        now = datetime.now(timezone.utc)
        pending_users.delete_many({"otp_expiry": {"$lt": now}})
        time.sleep(60)

threading.Thread(target=clean_expired_otps, daemon=True).start()

# -------------------------------
# AUTH ROUTES
# -------------------------------

@app.route("/signup", methods=["POST"])
def signup():
    data = request.get_json()
    name = data.get("name")
    contact = data.get("contact")
    password = data.get("password")

    if not name or not contact or not password:
        return jsonify({"error": "Missing required fields (name, contact, password)"}), 400

    if users.find_one({"contact": contact}):
        return jsonify({"error": "User already exists"}), 400

    otp = str(random.randint(100000, 999999))
    expiry = datetime.now(timezone.utc) + timedelta(minutes=5)
    hashed_pw = bcrypt.generate_password_hash(password).decode("utf-8")

    pending_users.delete_many({"contact": contact})

    pending_users.insert_one({
        "name": name,
        "contact": contact,
        "method": "email",
        "password": hashed_pw,
        "otp": otp,
        "otp_expiry": expiry,
        "login_otp": False
    })

    try:
        send_email_otp(contact, otp)
        return jsonify({"message": "Signup OTP sent to email"}), 200
    except Exception as e:
        pending_users.delete_one({"contact": contact})
        return jsonify({"error": f"Failed to send OTP: {str(e)}"}), 500

@app.route("/verify-otp", methods=["POST"])
def verify_otp():
    data = request.get_json()
    contact = data.get("contact")
    otp = data.get("otp")

    record = pending_users.find_one({"contact": contact, "login_otp": False})
    if not record:
        return jsonify({"error": "No OTP request found"}), 404

    now = datetime.now(timezone.utc)
    otp_expiry = record["otp_expiry"]
    if otp_expiry.tzinfo is None:
        otp_expiry = otp_expiry.replace(tzinfo=timezone.utc)

    if now > otp_expiry:
        pending_users.delete_one({"contact": contact})
        return jsonify({"error": "OTP expired. Please sign up again."}), 400

    if record["otp"] != otp:
        return jsonify({"error": "Invalid OTP"}), 400

    users.insert_one({
        "name": record["name"],
        "contact": record["contact"],
        "method": "email",
        "password": record["password"],
        "verified": True,
        "image": DEFAULT_PROFILE_PIC,
        "created_at": datetime.now(timezone.utc)
    })

    pending_users.delete_one({"contact": contact})
    return jsonify({"message": "Account verified successfully"}), 200

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json()
    contact = data.get("contact")
    password = data.get("password")

    user = users.find_one({"contact": contact})
    if not user:
        return jsonify({"error": "User not found"}), 404

    if bcrypt.check_password_hash(user["password"], password):
        return jsonify({
            "message": "Login successful",
            "name": user.get("name"),
            "contact": user.get("contact"),
            "userId": str(user.get("_id")),
            "image": user.get("image", DEFAULT_PROFILE_PIC)
        }), 200

    return jsonify({"error": "Invalid password"}), 400

@app.route("/send-login-otp", methods=["POST"])
def send_login_otp():
    data = request.get_json()
    contact = data.get("contact")

    user = users.find_one({"contact": contact})
    if not user:
        return jsonify({"error": "User not found"}), 404

    otp = str(random.randint(100000, 999999))
    expiry = datetime.now(timezone.utc) + timedelta(minutes=5)

    pending_users.delete_many({"contact": contact})
    pending_users.insert_one({
        "contact": contact,
        "otp": otp,
        "otp_expiry": expiry,
        "login_otp": True
    })

    try:
        send_email_otp(contact, otp)
        return jsonify({"message": "Login OTP sent to email"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to send OTP: {str(e)}"}), 500

@app.route("/verify-login-otp", methods=["POST"])
def verify_login_otp():
    data = request.get_json()
    contact = data.get("contact")
    otp = data.get("otp")

    record = pending_users.find_one({"contact": contact, "login_otp": True})
    if not record:
        return jsonify({"error": "No OTP login request found"}), 404

    now = datetime.now(timezone.utc)
    otp_expiry = record["otp_expiry"]
    if otp_expiry.tzinfo is None:
        otp_expiry = otp_expiry.replace(tzinfo=timezone.utc)

    if now > otp_expiry:
        pending_users.delete_one({"contact": contact})
        return jsonify({"error": "OTP expired"}), 400

    if record["otp"] != otp:
        return jsonify({"error": "Invalid OTP"}), 400

    user = users.find_one({"contact": contact})
    pending_users.delete_one({"contact": contact})

    return jsonify({
        "message": "Login successful",
        "name": user.get("name"),
        "contact": user.get("contact"),
        "userId": str(user.get("_id")),
        "image": user.get("image", DEFAULT_PROFILE_PIC)
    }), 200

# -------------------------------
# EXCEL / CATALOG ROUTES
# -------------------------------
_cache = {"data": None, "timestamp": 0}
CACHE_TTL = 10  # seconds

def load_items_by_category():
    global _cache
    if _cache["data"] and (time.time() - _cache["timestamp"] < CACHE_TTL):
        return _cache["data"]

    if not EXCEL_PATH.exists():
        print(f"Excel file not found at: {EXCEL_PATH}")
        return {}

    wb = openpyxl.load_workbook(EXCEL_PATH)
    categories = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        items = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            if len(row) < 2 or not row[0] or not row[1]:
                continue
            unique_id = f"{sheet_name}_{row[0]}"
            item = {
                "id": unique_id,
                "title": row[1],
                "author": row[2] if len(row) > 2 else "",
                "price": row[3] if len(row) > 3 else "",
                "image": row[4] if len(row) > 4 and row[4] else "",
                "category": sheet_name,
                "image2": row[6] if len(row) > 6 and row[6] else "",
                "image3": row[7] if len(row) > 7 and row[7] else "",
                "description": row[8] if len(row) > 8 and row[8] else "",
                "image4": row[9] if len(row) > 9 and row[9] else "",
                "image5": row[10] if len(row) > 10 and row[10] else "",
            }
            items.append(item)
        categories[sheet_name] = items

    _cache["data"] = categories
    _cache["timestamp"] = time.time()
    print(f"Loaded {len(categories)} categories from Excel.")
    return categories

@app.route("/api/category-list")
def get_category_list():
    data = load_items_by_category()
    category_list = []
    for name, items in data.items():
        thumbnail = ""
        for item in items:
            if item.get("image"):
                thumbnail = item["image"]
                break
        category_list.append({
            "name": name,
            "count": len(items),
            "thumbnail": thumbnail or "https://via.placeholder.com/150"
        })
    return jsonify(category_list)

@app.route("/api/categories")
def get_categories():
    return jsonify(load_items_by_category())

@app.route("/api/category/<path:category_name>")
def get_category(category_name):
    categories = load_items_by_category()
    category_name = unquote(category_name)
    items = categories.get(category_name)
    if not items:
        return jsonify({"error": f"No items found for category '{category_name}'"}), 404
    return jsonify(items)

@app.route("/api/item/<item_id>")
def get_item(item_id):
    categories = load_items_by_category()
    for items in categories.values():
        for item in items:
            if item["id"] == item_id:
                return jsonify(item)
    return jsonify({"error": f"Item '{item_id}' not found"}), 404

# -------------------------------
# Run
# -------------------------------
if __name__ == "__main__":
    print("Starting server")
    print(f"Excel path: {EXCEL_PATH}")
    app.run(host="0.0.0.0", port=5000)
